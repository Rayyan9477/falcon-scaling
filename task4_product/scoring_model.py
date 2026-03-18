"""
Co-Investment Intelligence Accelerator — Scoring Model
=======================================================
Reads the family office dataset and produces a scored Excel dashboard
that helps fund managers identify optimal co-investment partners.

Output: co_investment_accelerator.xlsx with four sheets:
  1. Co-Investment Dashboard  — scored data with conditional formatting
  2. By Region               — pivot-style regional summary
  3. By Sector               — sector breakdown with counts and avg scores
  4. Scoring Methodology      — explanation of scoring logic
"""

import math
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
INPUT_PATH = Path(r"d:\Repo\falcon-scaling\data\family_office_dataset.xlsx")
OUTPUT_PATH = Path(r"d:\Repo\falcon-scaling\task4_product\co_investment_accelerator.xlsx")

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(name="Calibri", size=10)

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(name="Calibri", size=10, color="006100")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
YELLOW_FONT = Font(name="Calibri", size=10, color="9C6500")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(name="Calibri", size=10, color="9C0006")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

METHODOLOGY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# ---------------------------------------------------------------------------
# Sector category mapping
# ---------------------------------------------------------------------------
SECTOR_CATEGORIES = {
    "Technology & Digital": [
        "Technology", "Tech", "Fintech", "AI", "Artificial Intelligence",
        "Software", "SaaS", "Digital", "Cybersecurity", "Data",
        "Machine Learning", "Blockchain", "Crypto", "Web3",
        "Deep Tech", "Semiconductors", "E-commerce", "Ecommerce",
        "Media", "Entertainment", "Telecom", "Telecommunications",
        "Digital Infrastructure", "EdTech", "Gaming",
    ],
    "Healthcare & Life Sciences": [
        "Healthcare", "Health", "Biotech", "Biotechnology", "Pharma",
        "Pharmaceutical", "Life Sciences", "Medical", "MedTech",
        "Health Tech", "HealthTech", "Genomics", "Diagnostics",
        "Wellness", "Mental Health",
    ],
    "Financial Services": [
        "Financial Services", "Finance", "Banking", "Insurance",
        "Fintech", "Payments", "Wealth Management", "Asset Management",
        "Private Equity", "Venture Capital", "Capital Markets",
    ],
    "Real Estate & Infrastructure": [
        "Real Estate", "Property", "Infrastructure", "Construction",
        "Hospitality", "Hotels", "Logistics", "Transportation",
        "Ports", "Airports", "Utilities",
    ],
    "Energy & Sustainability": [
        "Energy", "Renewable Energy", "Renewables", "Clean Energy",
        "Cleantech", "Climate", "Climate Tech", "Sustainability",
        "Oil", "Gas", "Oil & Gas", "Mining", "Natural Resources",
        "Solar", "Wind", "ESG", "Green", "AgriTech",
    ],
    "Consumer & Retail": [
        "Consumer", "Retail", "CPG", "Consumer Products",
        "Food", "Food & Beverage", "F&B", "Luxury", "Fashion",
        "Beauty", "D2C", "DTC", "Consumer Tech", "Restaurants",
        "Agriculture",
    ],
    "Industrials & Manufacturing": [
        "Industrials", "Industrial", "Manufacturing", "Aerospace",
        "Defense", "Automotive", "Chemicals", "Materials",
        "Supply Chain", "Shipping", "Aviation",
    ],
    "Education & Social Impact": [
        "Education", "EdTech", "Impact", "Social Impact",
        "Philanthropy", "Nonprofit",
    ],
}


# ---------------------------------------------------------------------------
# Scoring functions
# ---------------------------------------------------------------------------

def _safe_str(val) -> str:
    """Return a stripped lowercase string; handle NaN/None."""
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return ""
    return str(val).strip().lower()


def _has_value(val) -> bool:
    """Return True if the value is a non-empty, non-NaN string."""
    s = _safe_str(val)
    return s != "" and s != "nan" and s != "n/a" and s != "none"


def compute_coinvest_score(row: pd.Series) -> float:
    """
    Co-Investment Score (1-10) — measures how suitable and active
    a family office is as a co-investment partner.

    Components (max raw = 10):
      Co-Invest Frequency:  High=3, Medium=2, Low=1
      Direct Investment:    Yes=2, No=0
      LP Status:            Active LP=3, Selective LP=2, Minimal=1
      GP/Direct Status:     Active GP=2, Emerging GP=1.5, Direct Only=1, Selective=0.5
    """
    score = 0.0

    # Co-Invest Frequency (max 3)
    freq = _safe_str(row.get("Co-Invest Frequency"))
    if "high" in freq:
        score += 3
    elif "medium" in freq or "moderate" in freq:
        score += 2
    elif "low" in freq:
        score += 1

    # Direct Investment (max 2)
    direct = _safe_str(row.get("Direct Investment"))
    if "yes" in direct:
        score += 2

    # LP Status (max 3)
    lp = _safe_str(row.get("LP Status"))
    if "active" in lp:
        score += 3
    elif "selective" in lp:
        score += 2
    elif "minimal" in lp or "limited" in lp:
        score += 1

    # GP/Direct Status (max 2)
    gp = _safe_str(row.get("GP/Direct Status"))
    if "active gp" in gp:
        score += 2
    elif "emerging" in gp:
        score += 1.5
    elif "direct only" in gp or "direct" in gp:
        score += 1
    elif "selective" in gp:
        score += 0.5

    # Clamp to 1-10
    return max(1.0, min(10.0, score))


def compute_accessibility_score(row: pd.Series) -> float:
    """
    Accessibility Score (1-10) — measures how easy it is to reach
    and engage the family office.

    Components (max raw = 10):
      Data Confidence:     High=3, Medium=2, Low=1
      Website exists:      +1
      LinkedIn exists:     +1  (Primary DM LinkedIn or LinkedIn Company URL)
      Email Pattern exists: +1
      Contact Method:      Direct=2, Referral/Intro=1.5, Conference=1, Cold=0.5
      Phone exists:        +0.5
      Conference Attendance exists: +0.5
    """
    score = 0.0

    # Data Confidence (max 3)
    conf = _safe_str(row.get("Data Confidence"))
    if "high" in conf:
        score += 3
    elif "medium" in conf:
        score += 2
    elif "low" in conf:
        score += 1

    # Website (+1)
    if _has_value(row.get("Website")):
        score += 1

    # LinkedIn (+1) — check both DM LinkedIn and Company LinkedIn
    has_linkedin = (
        _has_value(row.get("Primary DM LinkedIn"))
        or _has_value(row.get("LinkedIn Company URL"))
    )
    if has_linkedin:
        score += 1

    # Email Pattern (+1)
    if _has_value(row.get("Email Pattern")):
        score += 1

    # Contact Method (max 2)
    method = _safe_str(row.get("Contact Method"))
    if "direct" in method:
        score += 2
    elif "referral" in method or "intro" in method:
        score += 1.5
    elif "conference" in method or "event" in method:
        score += 1
    elif method:
        score += 0.5

    # Phone (+0.5)
    if _has_value(row.get("Main Office Phone")):
        score += 0.5

    # Conference Attendance (+0.5)
    if _has_value(row.get("Conference Attendance")):
        score += 0.5

    return max(1.0, min(10.0, score))


def compute_outreach_readiness(coinvest: float, access: float) -> str:
    """Composite readiness label based on average of both scores."""
    composite = (coinvest + access) / 2.0
    if composite >= 7:
        return "Green"
    elif composite >= 4:
        return "Yellow"
    else:
        return "Red"


def classify_sectors(sector_str) -> str:
    """Map a comma-separated sector string into broad category labels."""
    if not _has_value(sector_str):
        return "Unclassified"

    raw = str(sector_str)
    matched = set()
    for category, keywords in SECTOR_CATEGORIES.items():
        for kw in keywords:
            if kw.lower() in raw.lower():
                matched.add(category)
                break

    return ", ".join(sorted(matched)) if matched else "Other"


# ---------------------------------------------------------------------------
# Excel writing helpers
# ---------------------------------------------------------------------------

def _style_header_row(ws, max_col: int, fill=None, font=None):
    """Apply header style to row 1."""
    fill = fill or HEADER_FILL
    font = font or HEADER_FONT
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_fit_columns(ws, min_width: int = 10, max_width: int = 40):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        width = max(min_width, min(max_width, max_len + 3))
        ws.column_dimensions[col_letter].width = width


def _apply_data_borders(ws, max_row: int, max_col: int, start_row: int = 2):
    """Apply thin borders and default font to data cells."""
    for row in range(start_row, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.font = DATA_FONT


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_dashboard_sheet(wb: Workbook, df: pd.DataFrame):
    """Sheet 1: Co-Investment Dashboard — main scored data."""
    ws = wb.active
    ws.title = "Co-Investment Dashboard"

    # Columns to include
    dashboard_cols = [
        "Family Office Name",
        "Type (SFO/MFO)",
        "Region",
        "HQ Country",
        "HQ City",
        "AUM ($B)",
        "Sector Focus",
        "Sector Category",
        "Co-Invest Frequency",
        "Direct Investment",
        "LP Status",
        "GP/Direct Status",
        "Investment Strategy",
        "Check Size Min ($M)",
        "Check Size Max ($M)",
        "Co-Investment Score",
        "Accessibility Score",
        "Outreach Readiness",
        "Co-Investor Relationships",
        "Fund Relationships",
        "Primary Decision Maker",
        "Primary DM Title",
        "Data Confidence",
        "Website",
    ]

    # Filter to columns that actually exist
    available_cols = [c for c in dashboard_cols if c in df.columns]

    # Write headers
    for ci, col_name in enumerate(available_cols, start=1):
        ws.cell(row=1, column=ci, value=col_name)

    _style_header_row(ws, len(available_cols))

    # Write data (sorted by Co-Investment Score descending)
    sorted_df = df.sort_values("Co-Investment Score", ascending=False).reset_index(drop=True)

    # Track column indices for conditional formatting
    coinvest_col = None
    access_col = None
    readiness_col = None

    for ci, col_name in enumerate(available_cols, start=1):
        if col_name == "Co-Investment Score":
            coinvest_col = ci
        elif col_name == "Accessibility Score":
            access_col = ci
        elif col_name == "Outreach Readiness":
            readiness_col = ci

    for ri, (_, row) in enumerate(sorted_df.iterrows(), start=2):
        for ci, col_name in enumerate(available_cols, start=1):
            val = row.get(col_name)
            if isinstance(val, float) and math.isnan(val):
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            # Number formatting for scores
            if col_name in ("Co-Investment Score", "Accessibility Score"):
                cell.number_format = "0.0"

            # Number formatting for AUM
            if col_name == "AUM ($B)":
                cell.number_format = "#,##0.0"

            # Number formatting for check sizes
            if col_name in ("Check Size Min ($M)", "Check Size Max ($M)"):
                cell.number_format = "#,##0"

    max_row = len(sorted_df) + 1
    max_col = len(available_cols)

    # Apply borders
    _apply_data_borders(ws, max_row, max_col)

    # Conditional formatting: Outreach Readiness (cell color)
    if readiness_col:
        for ri in range(2, max_row + 1):
            cell = ws.cell(row=ri, column=readiness_col)
            val = _safe_str(cell.value)
            if "green" in val:
                cell.fill = GREEN_FILL
                cell.font = GREEN_FONT
            elif "yellow" in val:
                cell.fill = YELLOW_FILL
                cell.font = YELLOW_FONT
            elif "red" in val:
                cell.fill = RED_FILL
                cell.font = RED_FONT

    # Conditional formatting: Co-Investment Score (gradient-like)
    if coinvest_col:
        for ri in range(2, max_row + 1):
            cell = ws.cell(row=ri, column=coinvest_col)
            try:
                v = float(cell.value)
            except (TypeError, ValueError):
                continue
            if v >= 8:
                cell.fill = GREEN_FILL
                cell.font = Font(name="Calibri", size=10, bold=True, color="006100")
            elif v >= 5:
                cell.fill = YELLOW_FILL
                cell.font = YELLOW_FONT
            else:
                cell.fill = RED_FILL
                cell.font = RED_FONT

    # Conditional formatting: Accessibility Score
    if access_col:
        for ri in range(2, max_row + 1):
            cell = ws.cell(row=ri, column=access_col)
            try:
                v = float(cell.value)
            except (TypeError, ValueError):
                continue
            if v >= 7:
                cell.fill = GREEN_FILL
                cell.font = Font(name="Calibri", size=10, bold=True, color="006100")
            elif v >= 4:
                cell.fill = YELLOW_FILL
                cell.font = YELLOW_FONT
            else:
                cell.fill = RED_FILL
                cell.font = RED_FONT

    # Freeze panes (freeze header row)
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    # Auto-fit columns
    _auto_fit_columns(ws)

    return ws


def build_region_sheet(wb: Workbook, df: pd.DataFrame):
    """Sheet 2: By Region — pivot-style regional summary."""
    ws = wb.create_sheet("By Region")

    # Aggregate by region
    region_agg = (
        df.groupby("Region")
        .agg(
            FO_Count=("Family Office Name", "count"),
            Avg_CoInvest_Score=("Co-Investment Score", "mean"),
            Avg_Accessibility_Score=("Accessibility Score", "mean"),
            Avg_AUM=("AUM ($B)", lambda x: pd.to_numeric(x, errors="coerce").mean()),
            Green_Count=("Outreach Readiness", lambda x: (x == "Green").sum()),
            Yellow_Count=("Outreach Readiness", lambda x: (x == "Yellow").sum()),
            Red_Count=("Outreach Readiness", lambda x: (x == "Red").sum()),
            Top_Sectors=("Sector Focus", lambda x: _top_sectors(x, n=3)),
        )
        .sort_values("Avg_CoInvest_Score", ascending=False)
        .reset_index()
    )

    headers = [
        "Region",
        "Family Office Count",
        "Avg Co-Investment Score",
        "Avg Accessibility Score",
        "Avg AUM ($B)",
        "Green (Ready)",
        "Yellow (Moderate)",
        "Red (Low)",
        "Top Sectors",
    ]

    for ci, h in enumerate(headers, start=1):
        ws.cell(row=1, column=ci, value=h)

    _style_header_row(ws, len(headers))

    for ri, (_, row) in enumerate(region_agg.iterrows(), start=2):
        ws.cell(row=ri, column=1, value=row["Region"])
        ws.cell(row=ri, column=2, value=row["FO_Count"])
        c3 = ws.cell(row=ri, column=3, value=round(row["Avg_CoInvest_Score"], 1))
        c3.number_format = "0.0"
        c4 = ws.cell(row=ri, column=4, value=round(row["Avg_Accessibility_Score"], 1))
        c4.number_format = "0.0"
        c5 = ws.cell(row=ri, column=5, value=round(row["Avg_AUM"], 1) if not math.isnan(row["Avg_AUM"]) else 0)
        c5.number_format = "#,##0.0"
        ws.cell(row=ri, column=6, value=row["Green_Count"])
        ws.cell(row=ri, column=7, value=row["Yellow_Count"])
        ws.cell(row=ri, column=8, value=row["Red_Count"])
        ws.cell(row=ri, column=9, value=row["Top_Sectors"])

    max_row = len(region_agg) + 1
    _apply_data_borders(ws, max_row, len(headers))

    # Color the readiness count columns
    for ri in range(2, max_row + 1):
        ws.cell(row=ri, column=6).fill = GREEN_FILL
        ws.cell(row=ri, column=6).font = GREEN_FONT
        ws.cell(row=ri, column=7).fill = YELLOW_FILL
        ws.cell(row=ri, column=7).font = YELLOW_FONT
        ws.cell(row=ri, column=8).fill = RED_FILL
        ws.cell(row=ri, column=8).font = RED_FONT

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max_row}"
    _auto_fit_columns(ws)


def build_sector_sheet(wb: Workbook, df: pd.DataFrame):
    """Sheet 3: By Sector — top sectors with FO counts and avg co-invest scores."""
    ws = wb.create_sheet("By Sector")

    # Explode sectors (each FO may have multiple)
    sector_rows = []
    for _, row in df.iterrows():
        raw = str(row.get("Sector Focus", ""))
        if not _has_value(raw):
            continue
        for s in raw.split(","):
            s = s.strip()
            if s:
                sector_rows.append({
                    "Sector": s,
                    "Co-Investment Score": row["Co-Investment Score"],
                    "Accessibility Score": row["Accessibility Score"],
                    "AUM ($B)": pd.to_numeric(row.get("AUM ($B)"), errors="coerce"),
                    "Outreach Readiness": row["Outreach Readiness"],
                })

    sector_df = pd.DataFrame(sector_rows)

    if sector_df.empty:
        ws.cell(row=1, column=1, value="No sector data available.")
        return

    sector_agg = (
        sector_df.groupby("Sector")
        .agg(
            FO_Count=("Sector", "count"),
            Avg_CoInvest_Score=("Co-Investment Score", "mean"),
            Avg_Accessibility_Score=("Accessibility Score", "mean"),
            Avg_AUM=("AUM ($B)", "mean"),
            Green_Count=("Outreach Readiness", lambda x: (x == "Green").sum()),
        )
        .sort_values("FO_Count", ascending=False)
        .reset_index()
    )

    headers = [
        "Sector",
        "Family Office Count",
        "Avg Co-Investment Score",
        "Avg Accessibility Score",
        "Avg AUM ($B)",
        "Green-Ready FOs",
    ]

    for ci, h in enumerate(headers, start=1):
        ws.cell(row=1, column=ci, value=h)

    _style_header_row(ws, len(headers))

    for ri, (_, row) in enumerate(sector_agg.iterrows(), start=2):
        ws.cell(row=ri, column=1, value=row["Sector"])
        ws.cell(row=ri, column=2, value=row["FO_Count"])
        c3 = ws.cell(row=ri, column=3, value=round(row["Avg_CoInvest_Score"], 1))
        c3.number_format = "0.0"
        c4 = ws.cell(row=ri, column=4, value=round(row["Avg_Accessibility_Score"], 1))
        c4.number_format = "0.0"
        aum_val = row["Avg_AUM"]
        c5 = ws.cell(row=ri, column=5, value=round(aum_val, 1) if not math.isnan(aum_val) else 0)
        c5.number_format = "#,##0.0"
        ws.cell(row=ri, column=6, value=row["Green_Count"])

    max_row = len(sector_agg) + 1
    _apply_data_borders(ws, max_row, len(headers))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max_row}"
    _auto_fit_columns(ws)


def build_methodology_sheet(wb: Workbook):
    """Sheet 4: Scoring Methodology — clear explanation of how scores work."""
    ws = wb.create_sheet("Scoring Methodology")

    content = [
        ("CO-INVESTMENT INTELLIGENCE ACCELERATOR", "Scoring Methodology", True),
        ("", "", False),
        ("1. CO-INVESTMENT SCORE (1-10)", "", True),
        ("Measures how suitable and active a family office is as a co-investment partner.", "", False),
        ("", "", False),
        ("Component", "Points", True),
        ("Co-Invest Frequency: High", "3 pts", False),
        ("Co-Invest Frequency: Medium", "2 pts", False),
        ("Co-Invest Frequency: Low", "1 pt", False),
        ("Direct Investment: Yes", "2 pts", False),
        ("Direct Investment: No", "0 pts", False),
        ("LP Status: Active LP", "3 pts", False),
        ("LP Status: Selective LP", "2 pts", False),
        ("LP Status: Minimal/Limited", "1 pt", False),
        ("GP/Direct Status: Active GP", "2 pts", False),
        ("GP/Direct Status: Emerging GP", "1.5 pts", False),
        ("GP/Direct Status: Direct Only", "1 pt", False),
        ("GP/Direct Status: Selective", "0.5 pts", False),
        ("Maximum possible raw score", "10 pts", False),
        ("", "", False),
        ("2. ACCESSIBILITY SCORE (1-10)", "", True),
        ("Measures how easy it is to reach and engage the family office.", "", False),
        ("", "", False),
        ("Component", "Points", True),
        ("Data Confidence: High", "3 pts", False),
        ("Data Confidence: Medium", "2 pts", False),
        ("Data Confidence: Low", "1 pt", False),
        ("Website exists", "+1 pt", False),
        ("LinkedIn profile exists (DM or Company)", "+1 pt", False),
        ("Email pattern available", "+1 pt", False),
        ("Contact Method: Direct", "+2 pts", False),
        ("Contact Method: Referral/Intro", "+1.5 pts", False),
        ("Contact Method: Conference/Event", "+1 pt", False),
        ("Contact Method: Other", "+0.5 pts", False),
        ("Phone number available", "+0.5 pts", False),
        ("Conference attendance listed", "+0.5 pts", False),
        ("Maximum possible raw score", "10 pts", False),
        ("", "", False),
        ("3. OUTREACH READINESS", "", True),
        ("Composite label based on the average of Co-Investment Score and Accessibility Score.", "", False),
        ("", "", False),
        ("Label", "Criteria", True),
        ("Green (Ready)", "Average score >= 7.0", False),
        ("Yellow (Moderate)", "Average score 4.0 - 6.9", False),
        ("Red (Low Priority)", "Average score < 4.0", False),
        ("", "", False),
        ("4. SECTOR MATCH CATEGORIES", "", True),
        ("Sectors from each family office are grouped into broad categories:", "", False),
        ("", "", False),
        ("Category", "Includes", True),
        ("Technology & Digital", "Tech, AI, Software, Fintech, SaaS, Digital, Cybersecurity, etc.", False),
        ("Healthcare & Life Sciences", "Healthcare, Biotech, Pharma, MedTech, Genomics, etc.", False),
        ("Financial Services", "Banking, Insurance, Fintech, Asset Management, etc.", False),
        ("Real Estate & Infrastructure", "Real Estate, Property, Infrastructure, Hospitality, Logistics, etc.", False),
        ("Energy & Sustainability", "Energy, Renewables, Clean Energy, Oil & Gas, Mining, etc.", False),
        ("Consumer & Retail", "Consumer, Retail, Food & Beverage, Luxury, Fashion, etc.", False),
        ("Industrials & Manufacturing", "Manufacturing, Aerospace, Automotive, Chemicals, etc.", False),
        ("Education & Social Impact", "Education, EdTech, Social Impact, Philanthropy, etc.", False),
    ]

    for ri, (col_a, col_b, is_header) in enumerate(content, start=1):
        cell_a = ws.cell(row=ri, column=1, value=col_a)
        cell_b = ws.cell(row=ri, column=2, value=col_b)

        if is_header:
            cell_a.font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
            cell_b.font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
            cell_a.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
            cell_b.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        else:
            cell_a.font = DATA_FONT
            cell_b.font = DATA_FONT

        cell_a.border = THIN_BORDER
        cell_b.border = THIN_BORDER
        cell_a.alignment = Alignment(vertical="center", wrap_text=True)
        cell_b.alignment = Alignment(vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 55


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _top_sectors(series: pd.Series, n: int = 3) -> str:
    """Extract the most common individual sectors from a series of comma-separated strings."""
    from collections import Counter
    counter = Counter()
    for val in series.dropna():
        for s in str(val).split(","):
            s = s.strip()
            if s:
                counter[s] += 1
    top = counter.most_common(n)
    return ", ".join(f"{s} ({c})" for s, c in top)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("Co-Investment Intelligence Accelerator — Build Pipeline")
    print("=" * 60)

    # 1. Load dataset
    print(f"\n[1/5] Loading dataset from {INPUT_PATH} ...")
    df = pd.read_excel(INPUT_PATH, sheet_name="Family Office Intelligence", engine="openpyxl")
    print(f"      Loaded {len(df)} family offices with {len(df.columns)} fields.")

    # 2. Compute scores
    print("[2/5] Computing Co-Investment Scores ...")
    df["Co-Investment Score"] = df.apply(compute_coinvest_score, axis=1)

    print("[3/5] Computing Accessibility Scores ...")
    df["Accessibility Score"] = df.apply(compute_accessibility_score, axis=1)

    print("[4/5] Computing Outreach Readiness & Sector Categories ...")
    df["Outreach Readiness"] = df.apply(
        lambda r: compute_outreach_readiness(r["Co-Investment Score"], r["Accessibility Score"]),
        axis=1,
    )
    df["Sector Category"] = df["Sector Focus"].apply(classify_sectors)

    # Summary stats
    green_count = (df["Outreach Readiness"] == "Green").sum()
    yellow_count = (df["Outreach Readiness"] == "Yellow").sum()
    red_count = (df["Outreach Readiness"] == "Red").sum()
    avg_coinvest = df["Co-Investment Score"].mean()
    avg_access = df["Accessibility Score"].mean()

    print(f"      Avg Co-Investment Score:  {avg_coinvest:.1f}")
    print(f"      Avg Accessibility Score:  {avg_access:.1f}")
    print(f"      Readiness: Green={green_count}, Yellow={yellow_count}, Red={red_count}")

    # 3. Build Excel workbook
    print(f"[5/5] Building Excel dashboard at {OUTPUT_PATH} ...")
    wb = Workbook()

    build_dashboard_sheet(wb, df)
    build_region_sheet(wb, df)
    build_sector_sheet(wb, df)
    build_methodology_sheet(wb)

    # Save
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_PATH))

    print(f"\n{'=' * 60}")
    print(f"SUCCESS: Dashboard saved to {OUTPUT_PATH}")
    print(f"  Sheets: {wb.sheetnames}")
    print(f"  Total family offices scored: {len(df)}")
    print(f"  Top 5 by Co-Investment Score:")
    top5 = df.nlargest(5, "Co-Investment Score")[
        ["Family Office Name", "Co-Investment Score", "Accessibility Score", "Outreach Readiness"]
    ]
    for _, r in top5.iterrows():
        print(f"    - {r['Family Office Name']}: CoInvest={r['Co-Investment Score']:.1f}, "
              f"Access={r['Accessibility Score']:.1f}, Readiness={r['Outreach Readiness']}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
